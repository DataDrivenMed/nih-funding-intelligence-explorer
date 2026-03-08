"""
NIH Funding Explorer — Step 1: Data Cleaning Pipeline
======================================================
Source: NIH Data Book Report ID 302 (Funding Patterns by IC)
Input:  NIH percentile.zip
Output:
  processed/nih_master_cleaned_data.csv
  logs/file_processing_log.csv
  logs/file_errors.csv

Inspection findings applied:
  - All 238 xlsx files share sheet name 'id', headers on row 5, data from row 6
  - 'D' suppression only in R56 Awards column → coerce to NaN
  - Percentile=0 exists in 2014 ALL NIH → drop
  - Institute name inconsistency: normalize "All NIH" → "ALL NIH"
  - Columns 7-15 are always empty → drop
"""

import sys
import io
import zipfile
import re
import logging
from pathlib import Path

import openpyxl
import pandas as pd
import numpy as np

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
ZIP_PATH = ROOT / "NIH percentile.zip"
PROCESSED_DIR = ROOT / "processed"
LOGS_DIR = ROOT / "logs"

PROCESSED_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Logging — console + file
# ---------------------------------------------------------------------------
log = logging.getLogger("nih_cleaner")
log.setLevel(logging.DEBUG)

_fmt = logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")

_stdout_utf8 = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
_ch = logging.StreamHandler(_stdout_utf8)
_ch.setLevel(logging.INFO)
_ch.setFormatter(_fmt)
log.addHandler(_ch)

_fh = logging.FileHandler(LOGS_DIR / "cleaning_run.log", mode="w", encoding="utf-8")
_fh.setLevel(logging.DEBUG)
_fh.setFormatter(_fmt)
log.addHandler(_fh)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
EXPECTED_SHEET = "id"
TITLE_ROW_IDX = 0        # 0-based index into rows list
METADATA_ROW_IDX = 2     # "Data for YYYY - INSTITUTE"
HEADER_ROW_IDX = 4       # column names
DATA_START_IDX = 5       # first data row

EXPECTED_COLS = ["Year", "Institute", "Percentile",
                 "R56 Awards", "R01-Equivalent Awards", "Not Awarded"]

# Columns we will keep after renaming
COL_RENAME = {
    "R56 Awards":            "R56_Awards",
    "R01-Equivalent Awards": "R01_Equivalent_Awards",
    "Not Awarded":           "Not_Awarded",
}

INSTITUTE_NORMALIZE = {
    "All NIH": "ALL NIH",
    "all nih": "ALL NIH",
}

TITLE_PATTERN = re.compile(r"Data for (\d{4})\s*-\s*(.+)", re.IGNORECASE)

# Percentile bands (closed on both ends for last bucket)
BAND_EDGES = [1, 6, 11, 16, 21, 26, float("inf")]
BAND_LABELS = ["1-5", "6-10", "11-15", "16-20", "21-25", "26+"]


# ---------------------------------------------------------------------------
# Helper: assign percentile band
# ---------------------------------------------------------------------------
def assign_band(p: float) -> str:
    for lo, hi, label in zip(BAND_EDGES[:-1], BAND_EDGES[1:], BAND_LABELS):
        if lo <= p < hi:
            return label
    return "26+"


# ---------------------------------------------------------------------------
# Core extractor: one xlsx file → DataFrame
# ---------------------------------------------------------------------------
def extract_file(zip_handle: zipfile.ZipFile, filename: str) -> dict:
    """
    Returns a dict with keys:
      'df'        – cleaned DataFrame (may be empty on error)
      'year'      – str or None
      'institute' – str or None
      'status'    – 'ok' | 'warning' | 'error'
      'notes'     – list of str
    """
    result = {"df": pd.DataFrame(), "year": None, "institute": None,
              "status": "ok", "notes": []}

    try:
        with zip_handle.open(filename) as fh:
            wb = openpyxl.load_workbook(io.BytesIO(fh.read()), data_only=True)
    except Exception as exc:
        result["status"] = "error"
        result["notes"].append(f"Cannot open workbook: {exc}")
        return result

    # --- Sheet check ---
    if EXPECTED_SHEET not in wb.sheetnames:
        result["status"] = "error"
        result["notes"].append(f"Sheet '{EXPECTED_SHEET}' not found; got {wb.sheetnames}")
        return result

    ws = wb[EXPECTED_SHEET]
    rows = [r for r in ws.iter_rows(values_only=True)]

    if len(rows) < DATA_START_IDX + 1:
        result["status"] = "error"
        result["notes"].append(f"Too few rows ({len(rows)}); expected >{DATA_START_IDX}")
        return result

    # --- Parse metadata row ---
    metadata_cell = rows[METADATA_ROW_IDX][0]
    meta_str = str(metadata_cell) if metadata_cell else ""
    m = TITLE_PATTERN.search(meta_str)
    if not m:
        result["status"] = "error"
        result["notes"].append(f"Cannot parse metadata row: '{meta_str}'")
        return result

    year = m.group(1).strip()
    institute = m.group(2).strip()
    result["year"] = year
    result["institute"] = institute

    # --- Validate header row ---
    header = [str(c).strip() if c is not None else "" for c in rows[HEADER_ROW_IDX]]
    actual_cols = [h for h in header[:6]]
    if actual_cols != EXPECTED_COLS:
        result["status"] = "warning"
        result["notes"].append(f"Unexpected headers: {actual_cols}")

    # --- Extract data rows ---
    data_rows = []
    for raw in rows[DATA_START_IDX:]:
        # Skip completely empty rows
        if all(v is None for v in raw):
            continue
        # Take only first 6 columns
        row_vals = list(raw[:6])
        # Pad if short
        while len(row_vals) < 6:
            row_vals.append(None)
        data_rows.append(row_vals)

    if not data_rows:
        result["status"] = "warning"
        result["notes"].append("No data rows found after header")
        return result

    df = pd.DataFrame(data_rows, columns=EXPECTED_COLS)

    # -----------------------------------------------------------------------
    # Cleaning steps
    # -----------------------------------------------------------------------

    # 1. Coerce 'D' suppression → NaN in R56 Awards
    d_mask = df["R56 Awards"] == "D"
    if d_mask.any():
        result["notes"].append(f"Suppressed 'D' in R56 Awards: {d_mask.sum()} rows → NaN")
        if result["status"] == "ok":
            result["status"] = "warning"
    df["R56 Awards"] = pd.to_numeric(df["R56 Awards"], errors="coerce")

    # 2. Coerce all numeric columns
    for col in ["Year", "Percentile", "R01-Equivalent Awards", "Not Awarded"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # 3. Drop percentile=0 (unscored/triaged artifact; only in 2014 ALL NIH)
    p0_rows = (df["Percentile"] == 0).sum()
    if p0_rows > 0:
        df = df[df["Percentile"] != 0].copy()
        result["notes"].append(f"Dropped {p0_rows} row(s) with Percentile=0")

    # 4. Drop rows where Percentile is NaN
    pnan = df["Percentile"].isna().sum()
    if pnan > 0:
        df = df.dropna(subset=["Percentile"]).copy()
        result["notes"].append(f"Dropped {pnan} row(s) with NaN Percentile")

    # 5. Normalize institute names
    df["Institute"] = df["Institute"].astype(str).str.strip()
    df["Institute"] = df["Institute"].replace(INSTITUTE_NORMALIZE)

    # 6. Override Year / Institute from metadata (more reliable than cell values)
    df["Year"] = int(year)
    df["Institute"] = institute if institute not in INSTITUTE_NORMALIZE else INSTITUTE_NORMALIZE[institute]

    # Re-apply normalization to the overridden value
    if df["Institute"].iloc[0] in INSTITUTE_NORMALIZE:
        df["Institute"] = INSTITUTE_NORMALIZE[df["Institute"].iloc[0]]

    # 7. Rename columns
    df = df.rename(columns=COL_RENAME)

    # 8. Derived variables
    #    Funded_Total = R56 + R01eq  (NaN propagates if R56 is suppressed)
    df["Funded_Total"] = df["R56_Awards"].fillna(0) + df["R01_Equivalent_Awards"].fillna(0)
    # Flag rows where R56 was suppressed so we know Funded_Total may be understated
    df["R56_Suppressed"] = d_mask.values[:len(df)] if len(d_mask) == len(df) else False

    # Recalculate: if R56 is NaN, use only R01eq for Funded_Total and flag
    r56_suppressed_mask = df["R56_Awards"].isna()
    df.loc[r56_suppressed_mask, "Funded_Total"] = df.loc[r56_suppressed_mask, "R01_Equivalent_Awards"]
    df["R56_Suppressed"] = r56_suppressed_mask

    df["Applications_Total"] = df["Funded_Total"] + df["Not_Awarded"].fillna(0)

    # Funding_Rate — avoid division by zero
    df["Funding_Rate"] = np.where(
        df["Applications_Total"] > 0,
        df["Funded_Total"] / df["Applications_Total"],
        np.nan
    )

    # 9. Percentile band
    df["Percentile_Band"] = df["Percentile"].apply(lambda p: assign_band(p) if pd.notna(p) else None)

    # 10. Final column order & types
    df["Year"] = df["Year"].astype(int)
    df["Percentile"] = df["Percentile"].astype(int)
    df = df[[
        "Year", "Institute", "Percentile",
        "R56_Awards", "R01_Equivalent_Awards", "Not_Awarded",
        "Funded_Total", "Applications_Total", "Funding_Rate",
        "R56_Suppressed", "Percentile_Band"
    ]]

    result["df"] = df
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    log.info("=" * 60)
    log.info("NIH Funding Explorer — Cleaning Pipeline")
    log.info("=" * 60)

    if not ZIP_PATH.exists():
        log.error(f"Zip file not found: {ZIP_PATH}")
        sys.exit(1)

    processing_log = []   # one row per file
    error_log = []        # one row per error/warning
    all_frames = []

    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        xlsx_files = sorted(
            [f for f in zf.namelist() if f.endswith(".xlsx")],
            key=lambda x: x.lower()
        )
        log.info(f"Found {len(xlsx_files)} xlsx files in archive")

        for i, fname in enumerate(xlsx_files, 1):
            short = Path(fname).name
            result = extract_file(zf, fname)

            row_count = len(result["df"])
            log_entry = {
                "file":      short,
                "year":      result["year"],
                "institute": result["institute"],
                "status":    result["status"],
                "rows":      row_count,
                "notes":     "; ".join(result["notes"]) if result["notes"] else "",
            }
            processing_log.append(log_entry)

            if result["status"] == "error":
                log.error(f"[{i:03d}/{len(xlsx_files)}] ERROR  {short}: {result['notes']}")
                error_log.append(log_entry)
            elif result["status"] == "warning":
                log.warning(f"[{i:03d}/{len(xlsx_files)}] WARN   {short} ({result['year']}, {result['institute']}): {result['notes']}")
                if any("Suppressed" not in n for n in result["notes"]):
                    error_log.append(log_entry)
            else:
                log.debug(f"[{i:03d}/{len(xlsx_files)}] ok     {short} ({result['year']}, {result['institute']}) → {row_count} rows")

            if not result["df"].empty:
                all_frames.append(result["df"])

    # -----------------------------------------------------------------------
    # Combine
    # -----------------------------------------------------------------------
    log.info("-" * 60)
    log.info(f"Combining {len(all_frames)} DataFrames …")

    master = pd.concat(all_frames, ignore_index=True)

    # Sort logically
    master = master.sort_values(["Year", "Institute", "Percentile"]).reset_index(drop=True)

    # Sanity: remove any remaining fully-duplicate rows
    dups = master.duplicated(subset=["Year", "Institute", "Percentile"]).sum()
    if dups:
        log.warning(f"Dropping {dups} duplicate Year+Institute+Percentile rows")
        master = master.drop_duplicates(subset=["Year", "Institute", "Percentile"]).reset_index(drop=True)

    # -----------------------------------------------------------------------
    # Save outputs
    # -----------------------------------------------------------------------
    out_master = PROCESSED_DIR / "nih_master_cleaned_data.csv"
    master.to_csv(out_master, index=False)
    log.info(f"Saved master dataset → {out_master}  ({len(master):,} rows)")

    log_df = pd.DataFrame(processing_log)
    out_log = LOGS_DIR / "file_processing_log.csv"
    log_df.to_csv(out_log, index=False)
    log.info(f"Saved processing log → {out_log}  ({len(log_df)} entries)")

    err_df = pd.DataFrame(error_log) if error_log else pd.DataFrame(columns=log_df.columns)
    out_err = LOGS_DIR / "file_errors.csv"
    err_df.to_csv(out_err, index=False)
    log.info(f"Saved error log      → {out_err}  ({len(err_df)} entries)")

    # -----------------------------------------------------------------------
    # Summary stats
    # -----------------------------------------------------------------------
    log.info("=" * 60)
    log.info("SUMMARY")
    log.info("=" * 60)
    log.info(f"  Files processed:       {len(xlsx_files)}")
    log.info(f"  Files OK:              {sum(1 for r in processing_log if r['status'] == 'ok')}")
    log.info(f"  Files with warnings:   {sum(1 for r in processing_log if r['status'] == 'warning')}")
    log.info(f"  Files with errors:     {sum(1 for r in processing_log if r['status'] == 'error')}")
    log.info(f"  Total rows (master):   {len(master):,}")
    log.info(f"  Years:                 {sorted(master['Year'].unique().tolist())}")
    log.info(f"  Institutes:            {sorted(master['Institute'].unique().tolist())}")
    log.info(f"  Percentile range:      {int(master['Percentile'].min())} – {int(master['Percentile'].max())}")
    log.info(f"  R56 suppressed rows:   {master['R56_Suppressed'].sum():,}")
    log.info(f"  Rows with NaN R56:     {master['R56_Awards'].isna().sum():,}")
    log.info(f"  Rows zero FundedTotal: {(master['Funded_Total'] == 0).sum():,}")
    log.info(f"  Funding_Rate range:    {master['Funding_Rate'].min():.4f} – {master['Funding_Rate'].max():.4f}")
    log.info("")
    log.info("  Rows per Year:")
    for yr, cnt in master.groupby("Year").size().items():
        log.info(f"    {yr}: {cnt:,}")
    log.info("")
    log.info("  Rows per Institute (ALL NIH only check):")
    all_nih = master[master["Institute"] == "ALL NIH"]
    log.info(f"    ALL NIH: {len(all_nih)} rows across {all_nih['Year'].nunique()} years")
    log.info("=" * 60)
    log.info("Cleaning pipeline complete.")

    return master


if __name__ == "__main__":
    main()
